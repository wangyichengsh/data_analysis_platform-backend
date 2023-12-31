import time, random
import os, sys
import json
import logging
import gc
import pickle
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from websocket import create_connection
import re
from jinja2 import Template

# pwd = os.path.dirname(os.path.realpath(__file__))
# sys.path.append(pwd+"/../../")
# os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'AnlsTool.settings')
import django
django.setup()

from .models import ParentTask, Version, Task, ExecFunction, Input, InputFileSheet, InputFileColumn, OutputSheet, OutputColumn, SqlCode, JobHistory
from utils import getConn,execcode, tranData, drop_temp_table
from django.contrib.auth import get_user_model
from django.db.models.query import QuerySet
from django.db.models import DateField, DateTimeField, sql,Max,Q
from django.db import close_old_connections, connection
from django.core.mail.message import EmailMessage
from django.core.mail import get_connection
from Feedback.models import FeedRecord

User = get_user_model()

logger = logging.getLogger('main.NormalTask_common')

channel_host = 'ws://129.211.26.73:8001'

##################################数据库连接#########################################
CONN_MAP = {}

class upfileException(Exception):
    def __init__(self, err="上传文件为空"):
        Exception.__init__(self, err)

def getOneConn(func_name):
    global CONN_MAP
    if CONN_MAP.get(func_name) == None:
       CONN_MAP[func_name] = getConn(func_name)
    

def getAllConn(input_sheet_detail,sql_detail):
    """
    根据上传文件配置，执行sql配置创建数据库连接
    若有python代码。则创建所有的数据库连接
    """
    for input_sheet in input_sheet_detail:
        exec_func = getExecFunc(input_sheet)
        getOneConn(exec_func.func_name)
    for sql in sql_detail:
        if not sql.if_mulit:
            exec_func = getExecFunc(sql)
            getOneConn(exec_func.func_name)
        else:
            exec_func = ExecFunction.objects.get(id=sql.from_conn)
            getOneConn(exec_func.func_name)
            exec_func = ExecFunction.objects.get(id=sql.to_conn)
            getOneConn(exec_func.func_name) 

#################################Python解析器#####################################
def PythonExec(code, res, input_dict,**conn):
    try: 
        logger.info(code)
        exec(code)
        return {'msg':'','res':[]}
    except Exception as e:
        return {'msg':str(e)}


##################################获取配置#########################################    
def getInputByTask(task):
    res = Input.objects.filter(task=task)
    return res
    
def getSqlByTask(task):
    res = SqlCode.objects.filter(task=task).order_by('sql_id')
    return res

def getInputSheetByTask(task):
    input_detail = getInputByTask(task)
    res = []
    for input in input_detail:
        if input.type == 'File' or input.type == 'List':
            sheet_set = InputFileSheet.objects.filter(input=input)
            for sheet in sheet_set:
                res.append(sheet)
    return res

def getExecFunc(odj):
    exec_func = odj.exec_id
    if isinstance(exec_func, int):
        exec_func = ExecFunction.objects.get(id=exec_func)
    if isinstance(exec_func, str):
        exec_func = ExecFunction.objects.get(id=int(exec_func))
    elif isinstance(exec_func, ExecFunction):
        exec_func = exec_func
    return exec_func
    
def getOutputSheetByTask(task):
    res = OutputSheet.objects.filter(task=task).order_by('sheet_output_id')
    return res
    
def getColumnBySheet(sheet):
    res = OutputColumn.objects.filter(sheet=sheet)
    return res
    
def copyDemandTask(from_id,to_id):
    """
    复制老需求下的最新版本任务到新需求
    """
    rawDemand = ParentTask.objects.get(id=from_id)
    newDemand = ParentTask.objects.get(id=to_id)
    rawDemand_task = {}
    v_d = {}
    for v in Version.objects.filter(parent_task=rawDemand):
        l = v_d.get(v.version_id, -1)
        if l < v.version_num:
            v_d[v.version_id] = v.version_num
            rawDemand_task[v.version_id] = {'task':v.son_task,'version':v}
    for id in rawDemand_task.keys():
        new_version = Version()
        old_version = rawDemand_task[id]['version']
        new_version.version_id = old_version.version_id
        new_version.version_num = old_version.version_num + 1
        new_version.parent_task = newDemand
        new_version.demand_seq = old_version.demand_seq
        new_task_id = copyTask(rawDemand_task[id]['task'].id)
        feed_list = FeedRecord.objects.filter(task_id=rawDemand_task[id]['task'].id)
        new_task = Task.objects.get(id=new_task_id)
        new_task.status = 'confirming'
        new_task.save()
        for feed in feed_list:
            feed.task_id = new_task
            feed.save()
        old_task = old_version.son_task
        old_task.if_valid = False
        old_task.save()
        new_version.son_task = Task.objects.get(id=new_task_id)
        new_version.if_exec = False
        new_version.save()
    return 

def acquire_tasklock():
    with open('tasklock','w') as f:
        f.write('lock')
    
def release_tasklock():
    with open('tasklock','w') as f:
        f.write('free')
        
def iftasklock():
    if not os.path.exists('tasklock'):
        release_tasklock()
    with open('tasklock','r') as f:
        s = f.read()
    if s == 'lock':
        return True
    else:
        return False
        
def copyTask(task_id):
    """
    复制一个任务，返回复制的任务id
    """
    while(iftasklock()):
        time.sleep(random.random())
    acquire_tasklock()
    try:
        new_task = Task.objects.get(id=task_id)
        old_task_id = new_task.id
        max_id = Task.objects.aggregate(Max('id')).get('id__max')
        new_task.id = max_id + 1
        # new_task.status = 'confirming'
        new_task.save()
        d = db2json(old_task_id,Pickle=False)
        if len(d['msg'])==0:
            stream = d['stream']
            xl2db(new_task.id, stream, mode='update',stream_type='json')
        else:
            logger.error('copyTask Error:'+d['msg'])
    except Exception as e:
        logger.error('copyTask Error:'+str(e))
    finally:
        release_tasklock()    
    return new_task.id
    
##################################执行操作相关#########################################    
def createExecHistory(task,input_param,user_id,uid):
    jh = JobHistory()
    jh.task = task
    user = User.objects.get(id=user_id)
    jh.user = user
    input_param_json = {}
    for key in input_param.keys():
        input_param_json[key] = str(input_param[key])
    jh.input_json = json.dumps(input_param_json, ensure_ascii=False)
    jh.sql_all = len(SqlCode.objects.filter(task=task))
    jh.save()
    return jh.id

def saveFileInput(input_detail, input_param, user_id, job_id, lock=None):
    """
    将上传文件放入本地static/input_file/{{user_id}}的目录下
    """
    filelist = []
    for input in input_detail:
        if input.type == 'File':
            upload_file = input_param[str(input.id)]
            if not os.path.exists('static'):
                os.mkdir('static')
            if not os.path.exists('static/input_file'):
                os.mkdir('static/input_file')
            if not os.path.exists('static/input_file/'+str(user_id)):
                os.mkdir('static/input_file/'+str(user_id))
            if upload_file=="undefined":
                raise upfileException()
            filename = upload_file.name.split('.')
            filename[-2] += str(int(time.time())*1000)
            filename = '.'.join(filename)
            with open('static/input_file/'+str(user_id)+'/'+filename,'wb') as f:
                f.write(upload_file.read())
            filelist.append(str(user_id)+'/'+filename)
    #记录到历史待实现
    if job_id != None:
        lock.acquire()
        jh = JobHistory.objects.get(id = job_id)
        jh.input_file = json.dumps(filelist, ensure_ascii=False)
        jh.renew_time = datetime.now()
        jh.save()
        lock.release()
    #记录到历史待实现
        
def getCreateSql(sheet,exec_func,orc_all=[]):
    """
    生成建表sql
    sheet 输入excel文件sheet页模型
    orc_all 所建的Orcale临时表list
    """
    if exec_func.db_type==0:
        table_name = str(sheet.name)
        c_temp_sql = 'DROP TABLE IF EXISTS '+ table_name +';\n'
        c_temp_sql += "CREATE LOCAL TEMPORARY TABLE " + table_name + '(\n'
    elif exec_func.db_type==1:
        ts_n = str(int(time.time()*1000))[-6:]
        table_name = str(sheet.name) + ts_n
        with open('OTT', mode='a') as f:
            f.write(str(exec_func.func_name)+':'+table_name+'\n')
        orc_all.append({'db':exec_func,'table':table_name,'raw_name':str(sheet.name)})
        c_temp_sql = 'CREATE GLOBAL TEMPORARY TABLE '+ table_name +'(\n'
    flag = 1
    col_set = InputFileColumn.objects.filter(sheet=sheet)
    for col in col_set:
        if flag == 1:
            flag = 0
        else:
            c_temp_sql += ','
        c_temp_sql += str(col.name) + '  ' + str(col.type)+'\n'
    if exec_func.db_type==0:
        # c_temp_sql += ')\nWITH (ORIENTATION = COLUMN)\nON COMMIT PRESERVE ROWS;'            
        c_temp_sql += ')\nON COMMIT PRESERVE ROWS;'            
    elif exec_func.db_type==1:
        c_temp_sql += ')\nON COMMIT PRESERVE ROWS;'
    return c_temp_sql,table_name,orc_all

def isAllNone(row,col_list, header_map):
    '''
    判断一行是否全为空
    全为空返回 True
    有不为空 返回 False
    '''
    res = True
    for col_name in col_list:
        if row[int(header_map[col_name])].value != None:
            return False
    return res
    
def rmNullCol(xlsx_data,col_list, header_map):
    '''
    删除全为空的行
    '''
    res = []
    for row in xlsx_data:
        if isAllNone(row, col_list, header_map):
            pass
        else:
            res.append(row)
    return res

def getInsertSql_simple(xlsx,sheet,table_name,exec_func):
    """ 
    生成插入Sql
    """
    xlsx_data = xlsx
    if len(xlsx)==0:
        return ''
    col_set = InputFileColumn.objects.filter(sheet=sheet)
    col_list = [str(col.name).strip().lower() for col in col_set]
    if exec_func.db_type==0:
        i_temp_sql = 'INSERT INTO '+ table_name + '('
        col_sql = ','.join(col_list)
        i_temp_sql += col_sql+') VALUES '            
        d_sql = ','.join(["('"+str(row) +"')" for row in xlsx_data])
        i_temp_sql += d_sql + ';'
    elif exec_func.db_type==1:
        batch_size = 100
        if len(xlsx_data)<batch_size:
            i_temp_sql = 'INSERT ALL \n'
            fix_sql = 'INTO '+ str(table_name) + '('+','.join(col_list)+') VALUES '
            i_temp_sql += ''.join([fix_sql+"('"+str(row).strip()+"')\n"  for row in xlsx_data])            
            i_temp_sql += 'select 1 from dual'
        else:
            i_temp_sql = ''
            for s in range(0, len(xlsx_data),batch_size):
                data = xlsx_data[s:s+batch_size]
                i_temp_sql += 'INSERT ALL \n'
                fix_sql = 'INTO '+ str(table_name) + '('+','.join(col_list)+') VALUES '
                i_temp_sql += ''.join([fix_sql+"('"+str(row).strip()+"')\n"  for row in data])
                i_temp_sql += 'select 1 from dual;'
    i_temp_sql = i_temp_sql.replace("'None'",'null')
    return i_temp_sql

def getInsertSql(ws,sheet,table_name,exec_func):
    """ 
    生成插入Sql
    """
    xlsx_data = list(ws)
    col_set = InputFileColumn.objects.filter(sheet=sheet)
    col_list = [str(col.name).strip().lower() for col in col_set]
    header_map = {}
    for i,header in enumerate(xlsx_data[0]):
        header_map[str(header.value).strip().lower()] = i
    if len(xlsx_data)>1:
        xlsx_data = xlsx_data[1:]
    elif len(xlsx_data)== 1 or len(xlsx_data)== 0:
        return ''
    xlsx_data = rmNullCol(xlsx_data, col_list, header_map)
    if exec_func.db_type==0:
        i_temp_sql = 'INSERT INTO '+ table_name + '('
        col_sql = ','.join(col_list)
        i_temp_sql += col_sql+') VALUES '            
        d_sql = ','.join(['('+','.join(["'"+str(row[int(header_map[col_name])].value).strip()+"'" for col_name in col_list]) +')' for row in xlsx_data])
        if len(d_sql)==0:
            return ''
        i_temp_sql += d_sql + ';'
    elif exec_func.db_type==1:
        batch_size = 100
        if len(xlsx_data)<batch_size:
            i_temp_sql = 'INSERT ALL \n'
            fix_sql = 'INTO '+ str(table_name) + '('+','.join(col_list)+') VALUES '
            i_temp_sql += ''.join([fix_sql+'('+','.join(["'"+str(row[int(header_map[col_name])].value).strip()+"'" if not isinstance(row[int(header_map[col_name])].value, datetime) else "to_date('"+str(row[int(header_map[col_name])].value)+"','YYYY-MM-DD HH24:MI:SS')" for col_name in col_list])+')\n' for row in xlsx_data])            
            i_temp_sql += 'select 1 from dual'
        else:
            i_temp_sql = ''
            for s in range(0, len(xlsx_data),batch_size):
                data = xlsx_data[s:s+batch_size]
                i_temp_sql += 'INSERT ALL \n'
                fix_sql = 'INTO '+ str(table_name) + '('+','.join(col_list)+') VALUES '
                i_temp_sql += ''.join([fix_sql+'('+','.join(["'"+str(row[int(header_map[col_name])].value).strip()+"'" if not isinstance(row[int(header_map[col_name])].value, datetime) else "to_date('"+str(row[int(header_map[col_name])].value)+"','YYYY-MM-DD HH24:MI:SS')" for col_name in col_list])+')\n' for row in data])
                i_temp_sql += 'select 1 from dual;'
    i_temp_sql = i_temp_sql.replace("'None'",'null')
    return i_temp_sql

def xl2twmpsql(input_id,file):
    """
    有上传excel文件生成对应sql
    """
    input = Input.objects.get(id=input_id)
    if not isinstance(file,str):
        wb = load_workbook(file)
        sheetname = wb.sheetnames
        dataType = 'excel'
    else:
        data = file.split('\n')
        if len(data) == 1:
            data = data[0].split(',')
        data = [i.rstrip('\r\n') for  i in data if len(i)>0]
        dataType = 'list'
    input_sheet = InputFileSheet.objects.filter(input=input).order_by('sheet_id')
    sql_text = ''
    for i,sheet in enumerate(input_sheet):
        exec_func = getExecFunc(sheet)
        sql_text +='--建表语句--\n' 
        c_temp_sql,table_name,orc_all = getCreateSql(sheet,exec_func,[])
        sql_text += c_temp_sql +'\n\n'
        if dataType == 'excel':
            ws = wb[sheetname[i]]
        if dataType == 'excel':
            i_temp_sql = getInsertSql(ws,sheet,table_name,exec_func)
        elif dataType == 'list':
            i_temp_sql = getInsertSql_simple(data,sheet,table_name,exec_func)
        sql_text += '--插入语句--\n'
        sql_text += i_temp_sql +'\n\n'
    return sql_text

def tranxl2db(input, input_param, job_id, lock=None):
    """
    生成临时表
    """
    global CONN_MAP
    res = ''
    input_sheet = InputFileSheet.objects.filter(input=input).order_by('sheet_id')
    if len(input_sheet)==0:
        res += 'file config is None!'
        print('file config is None!')
    else:
        if not isinstance(input_param[str(input.id)],str):
            wb = load_workbook(input_param[str(input.id)])
            sheetname = wb.sheetnames
            dataType = 'excel'
        else:
            data = input_param[str(input.id)].split('\n')
            if len(data) == 1:
                data = data[0].split(',')
            data = [i.rstrip('\r\n') for  i in data if len(i)>0]
            dataType = 'list'
        orc_all = []
        for i,sheet in enumerate(input_sheet):
            # 生成建表sql
            exec_func = getExecFunc(sheet)
            c_temp_sql,table_name,orc_all = getCreateSql(sheet,exec_func,orc_all)        
            if dataType == 'excel':
                ws = wb[sheetname[i]]
            if dataType == 'excel':
                i_temp_sql = getInsertSql(ws,sheet,table_name,exec_func)
            elif dataType == 'list':
                i_temp_sql = getInsertSql_simple(data,sheet,table_name,exec_func)
            c_res = execcode(sql=c_temp_sql,conf_name=str(exec_func.func_name),conn=CONN_MAP[exec_func.func_name])
            if len(i_temp_sql)>0:
                i_res = execcode(sql=i_temp_sql,conf_name=str(exec_func.func_name),conn=CONN_MAP[exec_func.func_name])
            else:
                i_res = {'msg':''}
            if len(c_res['msg'])!=0:
                res += 'Create Temp Table Fail:'+str(c_res['msg'])
                if job_id != None:
                    lock.acquire()
                    jh = JobHistory.objects.get(id = job_id) 
                    jh.exec_status = 4
                    jh.exec_info = str(c_res['msg'])
                    jh.renew_time = datetime.now()
                    jh.save()
                    lock.release()
            if len(i_res['msg'])!=0:
                res += 'Insert Into Temp Table Fail:'+str(i_res['msg'])
                if job_id != None:
                    jh = JobHistory.objects.get(id = job_id) 
                    if jh.exec_status != 4:
                        jh.exec_status = 4
                        jh.exec_info = str(i_res['msg'])
                        jh.renew_time = datetime.now()
                        jh.save()
    return res,orc_all 
        

def getuplowdict(d):
    res = {}
    for key in d.keys():
        res[key] = d[key]
        res[key.lower().strip()] = d[key]
        res[key.upper().strip()] = d[key]
    return res

def str2datetime(v):
    if len(v)==10:
        res = datetime.strptime(v,'%Y-%m-%d')
    elif len(v)==8:
        res = datetime.strptime(v,'%Y%m%d')
    elif len(v)==19:
        res = datetime.strptime(v,'%Y-%m-%d %H:%M:%S')
    elif len(v)==16:
        res = datetime.strptime(v,'%Y-%m-%dT%H:%M')
    elif len(v)==17:
        res = datetime.strptime(v,'%Y%m%d %H:%M:%S')
    elif len(v)==0:
        res = 'null'
    elif v == 'None':
        res = 'null'
    elif v == 'undefined':
        res = 'null'
    else:
        res = v
    return res

def getReplaceDict(input_param, input_detail, job_uuid):
    """
    根据输入字段类型生成占位符替换字典
    """
    replace_dict = {}
    replace_dict['uuid'] = "'"+str(job_uuid).strip()+"'"
    for input_id in input_param.keys():
        input_temp = input_detail.get(id=input_id)
        k = str(input_temp.replace_key).strip()
        v = str(input_param[input_id]).strip()
        if str(input_temp.type) == 'Date':            
            replace_dict[k] = str2datetime(v)           
        if str(input_temp.type) == 'DateTime':
            replace_dict[k] = str2datetime(v)
        elif str(input_temp.type) == 'String':
            replace_dict[k] = "'"+v+"'"
        elif str(input_temp.type) == 'Number':
            replace_dict[k] = v            
    return replace_dict

def rd2str(replace_dict, exec_func):
    res = {}
    for key in replace_dict.keys():
        if isinstance(replace_dict[key],datetime) or isinstance(replace_dict[key],date):
            if exec_func.db_type==0:
                res[key] = replace_dict[key].strftime("cast('%Y-%m-%d %H:%M:%S' as date)")
            elif exec_func.db_type==1:
                res[key] =  "to_date('"+str(replace_dict[key])+"','YYYY-MM-DD HH24:MI:SS')"
            elif exec_func.db_type==2:
                res[key] = "str_to_date('" + str(replace_dict[key]) +"','%Y-%m-%d')"
            else:
                res[key] = replace_dict[key].strftime("'%Y-%m-%d %H:%M:%S'")
        elif str(replace_dict[key])=='None':
            res[key] = 'null'
        elif len(replace_dict[key])==0:
            res[key] = 'null'
        else:
            res[key] = replace_dict[key]
    return res 

def getupdatecol(res):
    ud = {}
    for r in res['res']:
        d = r[-1]
        for key in d.keys():
            try:
                up[key] = datetme.strptime(str(d[key]), '%Y-%m-%d %H:%M:%S')
            except:
                ud[key] = d[key]
    return ud                

def getClearColumn(column_list):
    res = {}
    for column in column_list:
        res[str(column.replace_key).strip().lower()] = str(column.name).strip()
    return res

def makeCsv(res, sheet, file_list, sheet_list=[], index = -1):
    if sheet != None:
        fname = str(sheet.name)+str(int(round(time.time(),4)*10000))+'.csv'
    else:
        fname = str(int(round(time.time(),4)*10000))+'.csv'
    file_list.append(fname)
    if not os.path.exists('static'):
        os.mkdir('static')
    if not os.path.exists('static/output_file/'):
        os.mkdir('static/output_file/')
    csv_path = 'static/output_file/'+fname
    table_top = list(res[0].keys())
    column_list = getColumnBySheet(sheet)
    column_dict = getClearColumn(column_list)
    if len(sheet_list)>index:
        title = [column_dict.get(str(column_sql).strip().lower(), str(column_sql)) for column_sql in table_top]
    else:
        title = table_top
    with open(csv_path,mode='w',encoding='utf-8') as f:
        f.write(','.join(title)+'\n')
        for row in res:
            d = ''
            for i in row:
                if isinstance(row[i], float):
                    d += str(float(row[i]))+','
                elif isinstance(row[i],str):
                    temp_c = str(row[i]).replace(',','，').replace('\n',' ')
                    d+=temp_c+','
                else:
                    temp_c = str(row[i]).replace(',','，').replace('\n',' ')
                    d+=temp_c+','
            d = d[:-1]+'\n'
            f.write(d)
    return file_list
    
def makeExcel(task,l,max_row=1000000):
    file_list = []
    fname = str(task.title) + str(int(round(time.time(),4)*10000))+'.xlsx'
    file_list.append(fname)
    if not os.path.exists('static'):
        os.mkdir('static')
    if not os.path.exists('static/output_file/'):
        os.mkdir('static/output_file/')
    xlsx_path = 'static/output_file/'+fname
    wb = Workbook(write_only=True)
    sheet_list = getOutputSheetByTask(task)
    for index,table in enumerate(l):
        if len(sheet_list)>=(index+1):
            sheet = sheet_list[index]
            ws = wb.create_sheet(title=str(sheet.name))
        else:
            sheet = None
            ws = wb.create_sheet(title="Sheet"+str(index+1))
        if len(table['res'])>0 and len(table['res'])<=max_row:
            table_top = list(table['res'][0].keys())
            if len(sheet_list)>index:
                column_list = getColumnBySheet(sheet)
                column_dict = getClearColumn(column_list)
                title = [column_dict.get(str(column_sql).strip().lower(), str(column_sql)) for column_sql in table_top]
            else:
                title = table_top
            ws.append(title)
            for i in table['res']:
                r = [ i[j] for j in table_top]
                ws.append(r)
        elif len(table['res'])>max_row:
            file_list = makeCsv(table['res'], sheet, file_list, sheet_list=sheet_list, index=index)
            csv_name = file_list[-1]
            ws.append(['行数超过100w,以编码为gbk的csv导出。详见'+csv_name])
    wb.save(xlsx_path)
    wb.close()
    if 'ws' in dir():
        del ws,wb
    else:
        del wb
    gc.collect()
    return file_list

def replace_sql(sql,replace_dict,l='#{',r='}'):
    for key in replace_dict.keys():
        sql = re.sub(l+str(key).lower().strip()+r, str(replace_dict[key]), sql, flags=re.IGNORECASE)
        # sql = sql.replace(l+str(key).lower().strip()+r,str(replace_dict[key]))
        # print(temp_sql)
    return sql

def send_message(task, user_id, message):
    try:
        ws = create_connection(channel_host+'/ws/backend/'+str(user_id)+'/')
        msg = {'message':message}
        msg_str = json.dumps(msg)
        ws.send(msg_str)
        ws.close()
    except Exception as e:
        logger.error('websocket send error:'+ str(e))

def getDataType(data):
    data_sample = data[0]
    for col in data_sample:
        if data_sample[col] == None:
            for row in data:
                if row[col] != None:
                    data_sample[col] = row[col]
                    break
    return data_sample

def getResCreateSql(table_name, data):
    schema = django.conf.settings.RES_SCHEMA
    full_table_name = str(schema)+'.'+str(table_name)
    sql = 'CREATE UNLOGGED  TABLE '+ full_table_name +' (\n'
    if len(data)==0:
        return '', full_table_name
    data_sample = getDataType(data)
    flag = 1
    for col in data_sample.keys():
        data_type = {str:'varchar',float:'decimal',int:'decimal',datetime:'timestamp',date:'timestamp',type(None):'varchar'}[type(data_sample[col])]
        if flag == 1:
            flag = 0
        else:
            sql += ','
        sql += '"'+str(col)+'"' + '  ' + str(data_type)+'\n'
    sql += ')WITHOUT OIDS;'
    return sql,full_table_name

def getResInsertSql(table_name,data):
    sql = 'INSERT INTO '+ table_name + '('
    if len(data)>0:
        col_list = data[0].keys()
        col_list = ['"'+title +'"' for title in col_list]
        col_sql = ','.join(col_list)
        sql += col_sql+') VALUES '
        d_sql = ','.join(['('+','.join(["'"+str(row[k])+"'" for k in row.keys()]) +')' for row in data])
        sql += d_sql
        sql = sql.replace("'None'",'null')
        return sql
    else:
        return ''

def restoreRes(job_id, task, res, file_list):
    conn = connection
    for index,table in enumerate(res):
        table_name = 'res_'+ str(job_id) + '_' +str(index)
        create_sql,full_table_name = getResCreateSql(table_name, table['res'])
        execcode(create_sql, '结果导入postgre', conn)
        over_file_id = 1
        if (len(table['res'])<=1000000):
            batch_size = 10000
            if len(table['res'])<batch_size:
                insert_sql = getResInsertSql(full_table_name, table['res'])
                execcode(insert_sql, '结果导入postgre',conn)
            else:
                for s in range(0, len(table['res']),batch_size):
                    insert_sql = getResInsertSql(full_table_name, table['res'][s:s+batch_size])
                    execcode(insert_sql, '结果导入postgre', conn)
        else:
            try:
                sql = '''
                COPY {table} FROM STDIN 
                WITH CSV
                DELIMITER ','
                NULL 'None'
                HEADER 
                ENCODING 'utf-8'
                ;             
                '''.format(table=full_table_name)
                cur = conn.cursor()
                fname = file_list[over_file_id]
                over_file_id += 1
                file_path = 'static/output_file/'+fname
                with open(file_path, encoding='utf-8',mode='r') as f:
                    cur.copy_expert(sql, f)
                    conn.commit()
                cur.close()
            except Exception as e:
                logger.error('保存100w结果数据出错:'+str(e))



def getJobDone(task, input_param, user_id, job_uuid, job_id = None, lock=None):
    """
    主要函数
    用于执行代码，制作excel，返回json数据
    """
    try:
    # if 1 == 1:
        global CONN_MAP
        close_old_connections()
        # l用于存放结果的数据
        l = []
        # 保存上传文件到本地,replace_dict用于存放需要替换的占位符信息
        input_detail = getInputByTask(task)
        saveFileInput(input_detail, input_param, user_id, job_id, lock)
        replace_dict = getReplaceDict(input_param, input_detail, job_uuid)
        # 查看需要用到哪些数据库连接，并创建
        sql_detail = getSqlByTask(task)
        input_sheet_detail = getInputSheetByTask(task)
        getAllConn(input_sheet_detail,sql_detail)
        # 若有文件参数，写入对应临时表
        orc_all = []
        for input in input_detail:
            if input.type == 'File':
                res,orc_temp = tranxl2db(input,input_param,job_id, lock)
                orc_all.extend(orc_temp)
                # 出错则将报错信息加入结果数据
                if len(res) != 0:
                    l.append({'sql_id':'-1','res':[{'msg':str(res)}]})
            if input.type == 'List':
                res,orc_temp = tranxl2db(input,input_param,job_id, lock)
                orc_all.extend(orc_temp)
                # 出错则将报错信息加入结果数据
                if len(res) != 0:
                    l.append({'sql_id':'-1','res':[{'msg':str(res)}]})
        # oracle创建临时表用占位符表示
        for table in orc_all:
            replace_dict[table['raw_name']] = table['table']
        # 执行代码，若临时表插入失败则退出
        if job_id != None:
            jh = JobHistory.objects.get(id = job_id)
            if jh.exec_status == 4:
                return []
        for sql in sql_detail:
            d = {}
            d['sql_id'] = sql.sql_id
            exec_func = getExecFunc(sql)
            # 替换参数
            try:
                replace_dict = getuplowdict(replace_dict)
                temp_dict = rd2str(replace_dict, exec_func)
                if sql.replace_style == 0:
                    sql_str = str(sql.code) % temp_dict
                elif sql.replace_style == 1:
                    sql_str = replace_sql(str(sql.code),temp_dict,l='#{',r='}')
                elif sql.replace_style == 2:
                    sql_str = replace_sql(str(sql.code),temp_dict,l='&',r='')
            except Exception as e:
                logger.error('Replace key error:' + str(e))
                logger.error('Replace key error:' + str(sql.code))
                logger.error('Replace key error:' + str(temp_dict))
            # 执行代码
            try:
                template = Template(sql_str)
                sql_str = template.render(**temp_dict)
            except Exception as e:
                logger.error('jinja2 error:' + str(e)) 
            # print(sql_str)
            if not sql.if_mulit:
                exec_res = execcode(sql=sql_str,conf_name=str(exec_func.func_name),conn=CONN_MAP[exec_func.func_name])
                # 判断是否要替换参数
                if sql.file_type == 1 and len(exec_res['res'])>=1:
                    if len(exec_res['res'][-1]) >= 1:
                        replace_dict.update(getupdatecol(exec_res))
            else:
                tran_drop = {}
                from_conn = ExecFunction.objects.get(id=sql.from_conn)
                to_conn = ExecFunction.objects.get(id=sql.to_conn)
                if sql.if_temp:
                    import time
                    tran_table_name = str(sql.table_name)+str(int(time.time()*1000))[-6:]
                else:
                    tran_table_name = str(sql.table_name)
                exec_res,from_c,to_c = tranData(CONN_MAP[from_conn.func_name],CONN_MAP[to_conn.func_name],sql_str,tran_table_name, create = sql.if_create, to_conn_retry=to_conn.func_name, temp = sql.if_temp)
                replace_dict[str(sql.table_name).lower()] = tran_table_name
                replace_dict[str(sql.table_name).upper()] = tran_table_name
                replace_dict[str(sql.table_name)] = tran_table_name
                CONN_MAP[to_conn.func_name] = to_c
                CONN_MAP[from_conn.func_name] = from_c
                if sql.if_create:
                    temp = tran_drop.get(to_c,[])
                    temp.append(tran_table_name)
                    tran_drop[to_c] = temp
            if len(exec_res['msg'])>0:
                d['res'] = [{'msg':exec_res['msg'],'sql':exec_res['sql']}]
                l.append(d)
                if job_id != None:
                    lock.acquire()
                    jh = JobHistory.objects.get(id = job_id)
                    jh.exec_status = 2
                    jh.renew_time = datetime.now()
                    jh.save()
                    lock.release()
            else:
                # 判断是否要展示
                if sql.display == 0:
                    d['res'] = exec_res['res']
                    if len(d['res'])>0:
                        if type(d['res'][0]) == list:
                            for i in d['res']:
                                if len(i)>0 and 'msg' in i[0].keys():
                                    if job_id != None:
                                        lock.acquire()
                                        jh = JobHistory.objects.get(id = job_id)
                                        jh.exec_status = 2
                                        jh.renew_time = datetime.now()
                                        jh.save()
                                        lock.release()
                                temp = {'sql_id':d['sql_id'],'res':i}
                                l.append(temp)
                        else:
                            l.append(d)
                    else:
                        l.append(d)
                else:
                    d['res'] = exec_res['res']
                    if len(d['res'])>0:
                        if type(d['res'][0]) == list:
                            for i in d['res']:
                                if len(i)>0 and 'msg' in i[0].keys():
                                    if job_id != None:
                                        lock.acquire()
                                        jh = JobHistory.objects.get(id = job_id)
                                        jh.exec_status = 2
                                        jh.renew_time = datetime.now()
                                        jh.save()
                                        lock.release()
                                    temp = {'sql_id':d['sql_id'],'res':i}
                                    l.append(temp)
            if job_id != None:
                lock.acquire()
                jh = JobHistory.objects.get(id = job_id)
                jh.sql_done += 1
                jh.renew_time = datetime.now()
                jh.save()
                lock.release()
        # 制作excel或csv
        filelist = makeExcel(task,l)
        if job_id != None:
            lock.acquire()
            jh = JobHistory.objects.get(id = job_id)
            jh.res_file = json.dumps(filelist, ensure_ascii=False)
            if jh.exec_status == 0:
                jh.exec_status = 1
            jh.renew_time = datetime.now()
            jh.save()
            lock.release()
        # 结果传回前端
        pass
        # 删除创建的临时表
        for i in orc_all:
            drop_sql = 'truncate table '+str(i['table']) +';drop table '+str(i['table'])
            exec_func = i['db']
            d_res = execcode(sql=drop_sql,conf_name=str(exec_func.func_name),conn=CONN_MAP[exec_func.func_name])
        if 'tran_drop' in locals().keys():
            for conn in tran_drop.keys():
                drop_temp_table(tran_drop[conn],conn,drop= True)
        send_message(task, user_id,task.title+' finished') 
        # return filelist
    except Exception as e:
    # else:
        import time
        time.sleep(1)
        if job_id != None:
            lock.acquire()
            jh = JobHistory.objects.get(id = job_id)
            jh.exec_info = str(e)
            jh.exec_status = 3
            jh.renew_time = datetime.now()
            jh.save()
            lock.release()
        send_message(task, user_id,task.title+' failed:'+str(e)) 
        logger.error("后台错误:"+str(e))
    finally:
    # if 1==1:
        # 关闭连接
        try: 
            for k,conn in CONN_MAP.items():
                conn.close()
        except Exception as e:
            logger.error("关闭连接错误:"+str(e))
    # 结果存到本地数据库
    restoreRes(job_id, task, l, filelist)
    return 0

def getAllExecSql(task, input_param, job_uuid):
    """
    返回所有执行的sql代码，仅执行替换参数的sql
    """
    global CONN_MAP
    sql_text = ''
    orc_all = []
    input_detail = getInputByTask(task)
    replace_dict = getReplaceDict(input_param, input_detail, job_uuid)
    # 查看需要用到哪些数据库连接，并创建
    sql_detail = getSqlByTask(task)
    input_sheet_detail = getInputSheetByTask(task)
    getAllConn(input_sheet_detail,sql_detail)
    for input in input_detail:
        if input.type == 'File' or input.type == 'List':
            sql_text_one = xl2twmpsql(input.id,input_param[str(input.id)])
            sql_text += sql_text_one
            res,orc_temp = tranxl2db(input,input_param,None)
            orc_all.extend(orc_temp)
    for table in orc_all:
        replace_dict[table['raw_name']] = table['table']
    for sql in sql_detail:
        exec_func = getExecFunc(sql)
        try:
            replace_dict = getuplowdict(replace_dict)
            temp_dict = rd2str(replace_dict, exec_func)
            if sql.replace_style == 0:
                sql_str = str(sql.code) % temp_dict
            elif sql.replace_style == 1:
                sql_str = replace_sql(str(sql.code),temp_dict,l='#{',r='}')
            elif sql.replace_style == 2:
                sql_str = replace_sql(str(sql.code),temp_dict,l='&',r='')
        except Exception as e:
            sql_text += '\n'
            sql_text += '------------------替换参数错误sql---------------------\n'
            sql_text += sql.code
            sql_text += '------------------替换参数错误替换字典---------------------\n'
            sql_text += str(temp_dict) 
            sql_text += '\n\n'
        template = Template(sql_str)
        sql_str = template.render(**temp_dict)
        sql_text += '\n'
        sql_text += sql_str
        if sql.file_type == 1:
            exec_res = execcode(sql=sql_str,conf_name=str(exec_func.func_name),conn=CONN_MAP[exec_func.func_name])
            if len(exec_res['res'][-1]) >= 1:
                replace_dict.update(getupdatecol(exec_res))    
    # 删除创建的临时表
    for i in orc_all:
        drop_sql = 'truncate table '+str(i['table']) +';drop table '+str(i['table'])
        exec_func = i['db']
        d_res = execcode(sql=drop_sql,conf_name=str(exec_func.func_name),conn=CONN_MAP[exec_func.func_name])
    try: 
        for k,conn in CONN_MAP.items():
            conn.close()
    except Exception as e:
        logger.error("关闭连接错误:"+str(e))
    return sql_text

        
###########################################
##          后台配置更新删除相关         ##
###########################################         
def createversion(parent_id,param,create_user):
    """
    创建版本
    """
    res = ''
    try:
        parent_task = ParentTask.objects.get(id=parent_id)
        task = Task()
        task.title = str(param.get('title','')).strip()
        task.level = int(param.get('level','2'))
        task.status = 1
        task.if_valid = 1
        task.creator = create_user
        task.create_time = datetime.now()
        dev_id = int(param.get('developer'))
        task.developer = User.objects.get(id=dev_id)
        task.save()
        version = Version()
        if len(Version.objects.filter(parent_task=parent_task))>0:
            max_num = int(Version.objects.filter(parent_task=parent_task).order_by('-version_num')[0].version_num)
        else:
            max_num = 0
        version.version_num = max_num+1
        version.parent_task = parent_task
        version.son_task = task
        version.version_detail = str(param.get('detail',''))
        version.save()
        return res
    except Exception as e:
        res = str(e)
        return res

def renewversion(parent_id,param):
    """
    更新版本信息
    """
    res = ''
    try:
        parent_task = ParentTask.objects.get(id=parent_id)
        task_id = param.get('task_id')
        state = param.get('state')
        detail = param.get('detail')
        son_task = Task.objects.get(id=task_id)
        son_task.status = state
        son_task.save()
        version = Version.objects.filter(parent_task=parent_task).filter(son_task=son_task)[0]
        version.version_detail = detail
        print(version)
        print(detail)
        version.save()
        return res
    except Exception as e:
        res = str(e)
        return res

def deltaskdetail(task,sql=True):
    """
    删除子任务所有配置
    """
    input_s = Input.objects.filter(task=task)
    for i in input_s:
        i.delete()
    output_s = OutputSheet.objects.filter(task=task)
    for o in output_s:
        o.delete()
    sql = SqlCode.objects.filter(task=task)
    if sql == True:
        for s in sql:
            s.delete()
    return task

def qs2ws(qs, ws, title=None):
    if isinstance(qs ,QuerySet) and len(qs)>0:
        if title == None:
            k = list(qs[0].__dict__.keys())
            k_t = k
            k_k = k
        else:
            k = list(qs[0].__dict__.keys())
            k_t = [ title.get(key) for key in title.keys() if key in k ]
            k_k = [ key for key in title.keys() if key in k]
        ws.append(k_t)
        for row in qs:
            ws.append([row.__dict__[i] for i in k_k])
    if isinstance(qs ,list) and len(qs)>0:
        if title == None:
            k = list(qs[0].keys())
            k_t = k
            k_k = k
        else:
            k = list(qs[0].keys())
            k_t = [ title.get(key) for key in title.keys() if key in k ]
            k_k = [ key for key in title.keys() if key in k]
        ws.append(k_t)
        for row in qs:
            ws.append([row[i] for i in k_k])
        
def db2xl(task_id):
    try:
        task = Task.objects.get(id=task_id)
        wb = Workbook(write_only=True)
        ws1 = wb.create_sheet(title="输入配置")
        input_qs = Input.objects.filter(task=task)
        input_title = {'name':'输入参数','replace_key':'替换占位符','type':'输入类型','default_value':'默认值','detail':'说明'}
        qs2ws(input_qs, ws1, input_title)
        input_sheet_qs = InputFileSheet.objects.filter(name='==取空集筛选条件==')
        for input in input_qs:
            input_sheet_qs = input_sheet_qs.union(InputFileSheet.objects.filter(input = input))
        if len(input_sheet_qs)>0:
            ws_is = wb.create_sheet(title="输入文件配置")
            input_sheet_list = []
            for input_sheet in input_sheet_qs:
                d = input_sheet.__dict__
                d['input_name'] = input_sheet.input.name
                input_sheet_list.append(d)
            input_sheet_title = {'input_name':'输入参数','sheet_id':'输入表id','name':'对应临时表表名','exec_id_id':'对应数据库'}
            qs2ws(input_sheet_list, ws_is, input_sheet_title)
        if len(input_sheet_qs)>0:
            ws_col = wb.create_sheet(title="输入文件列名配置")
            input_file_column_qs = InputFileColumn.objects.filter(name='==取空集筛选条件==')
            for input_sheet in input_sheet_qs:
                input_file_column_qs =input_file_column_qs.union(InputFileColumn.objects.filter(sheet=input_sheet))
            input_file_column_list = []
            for input_file_column in input_file_column_qs:
                d = input_file_column.__dict__
                d['sheet_id'] = input_file_column.sheet.sheet_id
                input_file_column_list.append(d)
            input_col_title = {'sheet_id':'输入表id','name':'对应临时表列名','type':'字段类型'}
            qs2ws(input_file_column_list, ws_col, input_col_title)
        ws2 = wb.create_sheet(title="输出配置")
        output_sheet_qs = OutputSheet.objects.filter(task = task)
        sheetname_title = {'sheet_output_id':'输出表顺序','name':'输出表名称','detail':'输出表格算法描述'}
        qs2ws(output_sheet_qs, ws2, sheetname_title)
        ws3 = wb.create_sheet(title="输出字段信息")
        output_column_qs = OutputColumn.objects.filter(name="==取空集筛选条件==")
        for output_sheet in output_sheet_qs:
            output_column_qs = output_column_qs.union(OutputColumn.objects.filter(sheet=output_sheet).order_by('id'))
        output_column_qs = output_column_qs.order_by('id')    
        output_column_list = []
        for output_column in output_column_qs:
            d = output_column.__dict__
            d['output_id'] = output_column.sheet.sheet_output_id
            output_column_list.append(d)
        output_column_title = {'output_id':'输出表顺序','name':'输出字段名称','replace_key':'对应sql字段','detail':'说明'}
        qs2ws(output_column_list, ws3, output_column_title)
        ws4 = wb.create_sheet(title="SQL代码信息")
        sql_qs = SqlCode.objects.filter(task=task).order_by('sql_id')
        # sql_title = {'exec_id_id':'解析器id','file_type':'是否将结果作为占位符替换(0不替换,1替换)','display':'是否展示结果(0展示1不展示)','code':'sql代码','replace_style':'占位符替换风格'}
        sql_title = {'exec_id_id':'解析器id','file_type':'是否将结果作为占位符替换(0不替换,1替换)','display':'是否展示结果(0展示1不展示)','code':'sql代码','replace_style':'占位符替换风格','if_mulit':'是否跨库','from_conn':'导出解析器','to_conn':'导入解析器','table_name':'导入表名','if_create':'是否建表','if_temp':'是否建临时表'}
        qs2ws(sql_qs, ws4, sql_title)
        tmp = str(time.time())
        wb.save(tmp+".xlsx")
        wb.close()
        del ws1,ws2,ws3,ws4,wb
        with open(tmp+".xlsx",'rb') as f:
            stream = f.read()
        os.remove(tmp+".xlsx")    
        filename = task.title+'.xlsx'
        return {'stream':stream,'filename':filename,'msg':''}
    except Exception as e:
        return {'stream':None,'filename':None,'msg':str(e)}
     
def db2json(task_id,Pickle=True):
    try:
        task = Task.objects.get(id=task_id)
        wb = {}
        ws1 = []
        input_qs = Input.objects.filter(task=task)
        input_title = {'name':'输入参数','replace_key':'替换占位符','type':'输入类型','default_value':'默认值','detail':'说明'}
        qs2ws(input_qs, ws1, input_title)
        wb['输入配置'] = ws1
        input_sheet_qs = InputFileSheet.objects.filter(name='==取空集筛选条件==')
        for input in input_qs:
            input_sheet_qs = input_sheet_qs.union(InputFileSheet.objects.filter(input = input))
        if len(input_sheet_qs)>0:
            ws_is = []
            input_sheet_list = []
            for input_sheet in input_sheet_qs:
                d = input_sheet.__dict__
                d['input_name'] = input_sheet.input.name
                input_sheet_list.append(d)
            input_sheet_title = {'input_name':'输入参数','sheet_id':'输入表id','name':'对应临时表表名','exec_id_id':'对应数据库'}
            qs2ws(input_sheet_list, ws_is, input_sheet_title)
            wb['输入文件配置'] = ws_is
        if len(input_sheet_qs)>0:
            ws_col = []
            input_file_column_qs = InputFileColumn.objects.filter(name='==取空集筛选条件==')
            for input_sheet in input_sheet_qs:
                input_file_column_qs =input_file_column_qs.union(InputFileColumn.objects.filter(sheet=input_sheet))
            input_file_column_list = []
            for input_file_column in input_file_column_qs:
                d = input_file_column.__dict__
                d['sheet_id'] = input_file_column.sheet.sheet_id
                input_file_column_list.append(d)
            input_col_title = {'sheet_id':'输入表id','name':'对应临时表列名','type':'字段类型'}
            qs2ws(input_file_column_list, ws_col, input_col_title)
            wb['输入文件列名配置'] = ws_col
        ws2 = []
        output_sheet_qs = OutputSheet.objects.filter(task = task)
        sheetname_title = {'sheet_output_id':'输出表顺序','name':'输出表名称','detail':'输出表格算法描述'}
        qs2ws(output_sheet_qs, ws2, sheetname_title)
        wb['输出配置'] = ws2
        ws3 = []
        output_column_qs = OutputColumn.objects.filter(name="==取空集筛选条件==")
        for output_sheet in output_sheet_qs:
            output_column_qs = output_column_qs.union(OutputColumn.objects.filter(sheet=output_sheet).order_by('id'))
        output_column_qs = output_column_qs.order_by('id')
        output_column_list = []
        for output_column in output_column_qs:
            d = output_column.__dict__
            d['output_id'] = output_column.sheet.sheet_output_id
            output_column_list.append(d)
        output_column_title = {'output_id':'输出表顺序','name':'输出字段名称','replace_key':'对应sql字段','detail':'说明'}
        qs2ws(output_column_list, ws3, output_column_title)
        wb['输出字段信息'] = ws3
        ws4 = []
        sql_qs = SqlCode.objects.filter(task=task).order_by('sql_id')
        # sql_title = {'exec_id_id':'解析器id','file_type':'是否将结果作为占位符替换(0不替换,1替换)','display':'是否展示结果(0展示1不展示)','code':'sql代码','replace_style':'占位符替换风格'}
        sql_title = {'exec_id_id':'解析器id','file_type':'是否将结果作为占位符替换(0不替换,1替换)','display':'是否展示结果(0展示1不展示)','code':'sql代码','replace_style':'占位符替换风格','if_mulit':'是否跨库','from_conn':'导出解析器','to_conn':'导入解析器','table_name':'导入表名','if_create':'是否建表','if_temp':'是否建临时表'}
        qs2ws(sql_qs, ws4, sql_title)
        wb['SQL代码信息'] = ws4
        if Pickle == True:
            stream = pickle.dumps(wb)
            filename = task.title+'.pickle'
        else:   
            stream = wb
            filename = task.title+'.pickle'
        return {'stream':stream,'filename':filename,'msg':''}
    except Exception as e:
        return {'stream':None,'filename':None,'msg':str(e)}

def xl2output(task_id, stream,stream_type='workbook'):
    """
    从xlsx导入输出配置，删除原来的
    """
    res = ''
    try:
        task = Task.objects.get(id=task_id)
        output_s = OutputSheet.objects.filter(task=task)
        for o in output_s:
            o.delete()
        # wb = load_workbook(stream, read_only=True)
        if stream_type =='workbook':
            wb = load_workbook(stream)
            sheetlist = wb.sheetnames
        elif stream_type =='json':
            wb = stream
            sheetlist = wb.keys()
        elif stream_type =='pickle':
            wb = pickle.loads(stream)
            sheetlist = wb.keys()   
        for (index,sheet) in enumerate(sheetlist): 
            osheet = OutputSheet()
            osheet.task = task
            osheet.sheet_output_id = int(index)+1
            osheet.name = str(sheet).strip()
            osheet.save()
            ws = wb[sheet]
            row_iter = ws.rows
            if ws.max_row >= 2:
                try:
                    row1 = row_iter.send(None)
                except StopIteration:
                    row1 = []
                try:
                    row2_raw = row_iter.send(1)
                    row2 = [ i.value for i in row2_raw ]
                except StopIteration:
                    row2 = [''] * len(row1)
                try:
                    row3_raw = row_iter.send(2)
                    row3 = [ i.value for i in row3_raw ]
                except StopIteration:
                    row3 = [''] * len(row1)
            elif ws.max_row <=0:
                continue
            else:
                try:
                    row1 = row_iter.send(None)
                    row2 = [''] * len(row1)
                except StopIteration:
                    row1 = [] 
            for (i,c) in enumerate(row1):
                col = OutputColumn()
                col.name = str(c.value).strip()
                col.replace_key = row3[i] if row3[i] else ''
                col.detail = row2[i] if row2[i] else ''
                col.sheet = osheet
                col.save()
    except Exception as e:
        return str(e)
    else:
        return res
        
def xl2db(task_id, stream, mode='update', stream_type='workbook'):
    """
    从xlsx导入配置
    update：不删除原来配置
    renew：删除原来配置
    """
    res = ''
    task = Task.objects.get(id=task_id)
    # wb = load_workbook(stream, read_only=True)
    if stream_type =='workbook':
        wb = load_workbook(stream)
        sheetlist = wb.sheetnames
    elif stream_type =='json':
        wb = stream
        sheetlist = wb.keys()
    elif stream_type =='pickle':
        wb = pickle.loads(stream)
        sheetlist = wb.keys()   
    if mode == 'renew':
        deltaskdetail(task,sql=False)
    if '输入配置' in sheetlist:
        ws = wb['输入配置']
        if stream_type == 'workbook':
            iter = ws.values
        else:
            iter = ws
        for i,row in enumerate(iter):
            if i == 0:
                continue
            try:
                input = Input()
                input.task = task
                input.name = str(row[0]).strip()
                input.type = str(row[2]).strip()
                input.detail = str(row[4]).strip() if row[4] else ''
                input.default_value = str(row[3]).strip()
                input.replace_key = str(row[1]).strip()
                input.save()
            except Exception as e:
                logger.error('输入配置更新失败:'+str(e))
                res += '输入配置更新失败:'+str(e)+'\n'
    iss = {}
    if '输入文件配置' in sheetlist:
        ws = wb['输入文件配置']
        if stream_type == 'workbook':
            iter = ws.values
        else:
            iter = ws
        for i,row in enumerate(iter):
            if i == 0:
                continue
            try:
                input_sheet = InputFileSheet()
                input_s = Input.objects.filter(task=task).filter(name=str(row[0]).strip())[0]
                input_sheet.input = input_s
                input_sheet.sheet_id = int(row[1])
                input_sheet.name = str(row[2]).strip()
                ef = ExecFunction.objects.get(exec_id=int(row[3]))
                input_sheet.exec_id = ef
                input_sheet.save()
                iss[int(row[1])] = input_sheet
            except Exception as e:
                logger.error('输入文件配置更新失败:'+str(e))
                res += '输入文件配置更新失败:'+str(e)+'\n'
    if ('输入文件列名配置' in sheetlist) and len(iss)>0:
        ws = wb['输入文件列名配置']
        if stream_type == 'workbook':
            iter = ws.values
        else:
            iter = ws
        for i,row in enumerate(iter):
            if i == 0:
                continue
            try:
                input_col = InputFileColumn()
                input_col.sheet = iss[int(row[0])]
                input_col.name = str(row[1]).strip()
                input_col.type = str(row[2]).strip()
                input_col.save()
            except Exception as e:
                logger.error('输入文件列名配置更新失败:'+str(e))
                res += '输入文件列名配置更新失败:'+str(e)+'\n'
    oss = {}
    if '输出配置' in sheetlist:
        ws = wb['输出配置']
        if stream_type == 'workbook':
            iter = ws.values
        else:
            iter = ws
        for i,row in enumerate(iter):
            if i == 0:
                continue
            try:
                osheet = OutputSheet()
                osheet.task = task
                osheet.sheet_output_id = int(row[0])
                osheet.name = str(row[1]).strip()
                if len(row)>=3:
                    osheet.detail = str(row[2]).strip()
                osheet.save()
                oss[int(row[0])] = osheet
            except Exception as e:
                logger.error('输出配置更新失败:'+str(e))
                res += '输出配置更新失败:'+str(e)+'\n'
    if ('输出字段信息' in sheetlist) and len(oss)>0:
        ws = wb['输出字段信息']
        if stream_type == 'workbook':
            iter = ws.values
        else:
            iter = ws
        for i,row in enumerate(iter):
            if i == 0:
                continue
            try:
                col = OutputColumn()
                o_sheet = OutputSheet.objects.filter(task=task).filter(sheet_output_id=int(row[0]))[0]
                col.name = str(row[1]).strip()
                col.replace_key = str(row[2]).strip()
                col.detail = str(row[3]).strip() if row[3] else ''
                col.sheet = oss[int(row[0])]
                col.save()
            except Exception as e:
                logger.error('输出字段信息更新失败:'+str(e))
                res += '输出字段信息更新失败:'+str(e)+'\n'
    if 'SQL代码信息' in sheetlist:
    # if 1==2:
        ws = wb['SQL代码信息']
        if stream_type == 'workbook':
            iter = ws.values
            return res
        else:
            iter = ws
        for i,row in enumerate(iter):
            if i == 0:
                continue
            try:
                code = SqlCode()
                code.task = task
                ef = ExecFunction.objects.get(exec_id=row[0])
                code.exec_id = ef
                code.file_type = int(row[1])
                code.display = int(row[2])
                code.code = str(row[3])
                code.replace_style = int(row[4]) if len(row)>=5 else 0
                code.sql_id = i
                code.if_mulit = row[5] if len(row)>=6 else False
                if len(row)>=7:
                    code.from_conn = int(row[6]) if row[6]!=None  else 0
                if len(row)>=8:
                    code.to_conn = int(row[7]) if row[7]!=None  else 0
                code.table_name = str(row[8]) if len(row)>=9 else ''
                code.if_create = row[9] if len(row)>=10 else False
                code.if_temp = row[10] if len(row)>=11 else False
                code.save()
            except Exception as e:
                logger.error('SQL代码信息更新失败:'+str(e))
                res += 'SQL代码信息更新失败:'+str(e)+'\n'
    return res        

def send_mail(**kargs):
    connection = get_connection()
    mail = EmailMessage(from_email='jyjg_tech@ssein.com.cn', connection=connection ,**kargs)
    mail.send()
    connection.close()
    return True
